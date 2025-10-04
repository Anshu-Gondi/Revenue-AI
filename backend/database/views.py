import base64
import io
import zipfile
from django.shortcuts import render
from django.core.paginator import Paginator
from django.http import FileResponse, JsonResponse, HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import SavedResult, ContactMessage
import polars as pl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import pandas as pd
import json
import csv

# Create your views here.

""" Predication Page views """

# ─── Helper: Flatten nested JSON ─────────────────────────────


def flatten_json(data):
    def _flatten(d, parent_key="", sep="_"):
        items = []
        if isinstance(d, dict):
            if not d:
                items.append((parent_key, None))
            else:
                for k, v in d.items():
                    new_key = f"{parent_key}{sep}{k}" if parent_key else k
                    items.extend(_flatten(v, new_key, sep=sep).items())
        elif isinstance(d, list):
            if not d:
                items.append((parent_key, []))
            else:
                for i, v in enumerate(d):
                    new_key = f"{parent_key}{sep}{i}" if parent_key else str(i)
                    items.extend(_flatten(v, new_key, sep=sep).items())
        else:
            items.append((parent_key, d))
        return dict(items)
    return _flatten(data)


# ─── Unified Download View ───────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_saved_result(request, pk, file_type='json'):
    obj = SavedResult.objects.filter(pk=pk, owner=request.user).first()
    if not obj:
        return Response({'error': 'Not found or not yours'}, status=404)

    # Build combined JSON
    combined_json = {
        "id": obj.id,
        "file_name": obj.file_name,
        "notes": obj.notes,
        "uploaded_at": obj.uploaded_at.isoformat(),
        "inferred_target": obj.inferred_target,
        "model_name": obj.model_name,
        "data_shape": obj.data_shape,
        "eda_result": obj.eda_result or {},
        "model_result": obj.model_result or {},
    }

    file_type = file_type.lower()

    # ─── JSON ───────────────────────────────
    if file_type == 'json':
        return JsonResponse(combined_json, safe=False)

    # ─── CSV ────────────────────────────────
    if file_type == 'csv':
        flat_data = flatten_json(combined_json)
        df = pl.DataFrame([flat_data])
        buffer = io.StringIO()
        df.write_csv(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{obj.file_name}.csv"'
        return response

    # ─── Excel ──────────────────────────────
    if file_type in ['xlsx', 'excel']:
        flat_data = flatten_json(combined_json)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Saved Result"

        # Write key-value pairs (excluding graphs first)
        row = 1
        for key, val in flat_data.items():
            if "graphs" in key:
                continue  # handle separately
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(val))
            row += 1

        # Handle graph images (EDA + Model)
        graphs = {}

        # Include EDA graphs
        graphs.update(combined_json.get("eda_result", {}).get("graphs", {}))

        # Include model visualizations
        model_result = combined_json.get("model_result", {})
        if "graphs" in model_result:
            graphs.update(model_result["graphs"])
        if "diagnostic_graphs" in model_result:
            graphs.update(model_result["diagnostic_graphs"])
        if model_result.get("forecast_plot_base64"):
            graphs["forecast_plot"] = model_result["forecast_plot_base64"]

        if graphs:
            ws = wb.create_sheet("Graphs")
            row = 1

            # Add section headers
            ws.cell(row=row, column=1, value="Graph Visualizations")
            row += 2

            for name, b64 in graphs.items():
                try:
                    img_data = base64.b64decode(b64)
                    pil_img = PILImage.open(io.BytesIO(img_data)).convert("RGB")

                    img_buffer = io.BytesIO()
                    pil_img.save(img_buffer, format="PNG")
                    img_buffer.seek(0)

                    xl_img = XLImage(img_buffer)
                    ws.add_image(xl_img, f"A{row}")

                    # Caption below image
                    ws.cell(row=row + 15, column=1, value=name)
                    row += 30
                except Exception:
                    continue

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{obj.file_name}.xlsx"'
        return response

    # ─── Graph export (PNG or ZIP) ──────────
    if file_type == 'png':
        graphs = {}
        graphs.update(combined_json.get("eda_result", {}).get("graphs", {}))

        model_result = combined_json.get("model_result", {})
        if "graphs" in model_result:
            graphs.update(model_result["graphs"])
        if "diagnostic_graphs" in model_result:
            graphs.update(model_result["diagnostic_graphs"])
        if model_result.get("forecast_plot_base64"):
            graphs["forecast_plot"] = model_result["forecast_plot_base64"]

        if not graphs:
            return Response({"error": "No graphs available for PNG export"}, status=400)

        # Single graph → return one PNG
        if len(graphs) == 1:
            name, b64 = list(graphs.items())[0]
            try:
                image_data = base64.b64decode(b64)
                return HttpResponse(
                    image_data,
                    content_type="image/png",
                    headers={
                        "Content-Disposition": f'attachment; filename="{name}.png"'
                    }
                )
            except Exception:
                return Response({"error": "Failed to decode graph image"}, status=500)

        # Multiple graphs → return ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for name, b64 in graphs.items():
                try:
                    img_data = base64.b64decode(b64)
                    zipf.writestr(f"{name}.png", img_data)
                except Exception:
                    continue
        zip_buffer.seek(0)
        return FileResponse(zip_buffer, as_attachment=True, filename=f"{obj.file_name}_graphs.zip")

    # ─── Invalid Type ───────────────────────
    return JsonResponse(
        {"error": "Invalid file_type, use 'json', 'csv', 'xlsx', or 'png'."}, status=400)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_saved_results(request):
    page_size = int(request.GET.get('page_size', 10))
    page = int(request.GET.get('page', 1))

    qs = SavedResult.objects.filter(
        owner=request.user).order_by('-uploaded_at')
    paginator = Paginator(qs, page_size)
    page_obj = paginator.page(page)

    data = []
    for r in page_obj.object_list:
        data.append({
            "id":             r.id,
            "file_name":      r.file_name,
            "inferred_target": r.inferred_target,
            "data_shape":     r.data_shape,
            "result_json": {
                "eda_result":   r.eda_result or {},
                "model_result": r.model_result or {},
            },
            "model_name": r.model_name,
            "notes":      r.notes,
            "created_at": r.uploaded_at.isoformat(),
        })

    return JsonResponse({
        "results":      data,
        "total_pages":  paginator.num_pages,
        "current_page": page
    }, safe=False)


@api_view(['POST'])
def save_result_view(request):
    try:
        file_name = request.data.get("file_name", "unknown_file.csv")
        inferred_target = request.data.get("inferred_target")
        data_shape = request.data.get("data_shape")
        result_json = request.data.get("result_json")
        model_name = request.data.get("model_name")
        notes = request.data.get("notes", "")

        if isinstance(result_json, str):
            result_json = json.loads(result_json)

        # Decide EDA vs Model payload
        eda_payload = {}
        model_payload = {}
        if isinstance(result_json, dict):
            if "graphs" in result_json or "columns" in result_json:
                eda_payload = result_json
            elif "target_column" in result_json or "rmse" in result_json:
                model_payload = result_json
            else:
                eda_payload = result_json
        else:
            eda_payload = {}

        # ─── Check if this model was already run for this dataset ───
        existing_obj = SavedResult.objects.filter(
            owner=request.user,
            file_name=file_name,
            inferred_target=inferred_target,
            model_name=model_name
        ).first()

        if existing_obj:
            # Update the existing entry
            existing_obj.data_shape = data_shape
            existing_obj.eda_result = eda_payload
            existing_obj.model_result = model_payload
            existing_obj.notes = notes
            existing_obj.save()
            return Response({
                "message": f"Existing result for model '{model_name}' updated successfully.",
                "id": existing_obj.id
            })

        # ─── Save new result if it doesn't exist ──────────────────
        saved = SavedResult.objects.create(
            owner=request.user,
            file_name=file_name,
            inferred_target=inferred_target,
            data_shape=data_shape,
            eda_result=eda_payload,
            model_result=model_payload,
            model_name=model_name,
            notes=notes,
        )

        return Response({"message": "Result saved successfully.", "id": saved.id})

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_saved_result(request, pk):
    obj = SavedResult.objects.filter(pk=pk, owner=request.user).first()
    if not obj:
        return Response({'error': 'Not found or not yours'}, status=404)

    obj.notes = request.data.get('notes', obj.notes)
    obj.save()
    return Response({'message': 'Updated successfully'})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_result_view(request, result_id):
    obj = SavedResult.objects.filter(id=result_id, owner=request.user).first()
    if not obj:
        return Response({'error': 'Not found or not yours'}, status=404)
    obj.delete()
    return Response({"message": "Deleted"})


""" Contact Page views """


@api_view(['POST'])
def contact_form_view(request):
    try:
        name = request.data.get('name')
        email = request.data.get('email')
        message = request.data.get('message')

        # save to DB
        ContactMessage.objects.create(
            name=name,
            email=email,
            message=message
        )

        return Response({"message": "Message received successfully!"})
    except Exception as e:
        return Response({"error": str(e)}, status=500)
